import json
import string
import random
import os
from collections import defaultdict

import openpyxl
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from haystack.query import SearchQuerySet
from .models import Student, StudentsProgress, ClassModel, Subject, User, CollageGroups, WorkedOutLogs, Teacher, \
    Curator, Subgroup, TypeForLesson, ThemeForLesson
from leave_journals.models import OperationLog
from django.views.generic import View, DeleteView
from datetime import datetime, date
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from json import dumps
import urllib.parse
from django.db import transaction
from django.db.models import Q


class ItemJournalView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, item, group_code, subgroup_name, *args, **kwargs):
        month = int(request.GET.get('month', datetime.now().month))
        year = datetime.now().year
        group = get_object_or_404(CollageGroups, group_code=group_code)
        subject = get_object_or_404(Subject, item_name=item)
        curator = Curator.objects.filter(group=group).first()
        teachers_for_subject = subject.get_teachers_for_group(group)
        all_teachers = []
        if teachers_for_subject:
            all_teachers = Teacher.objects.values_list("teacher_full_name", flat=True).exclude(
                teacher_full_name=teachers_for_subject[0].teacher_full_name).order_by("teacher_full_name")

        subgroup_obj = Subgroup.objects.filter(name=subgroup_name, parent_group=group).first()
        if subgroup_name == "all":
            students_for_item = Student.objects.filter(group=group).order_by(
                'student_full_name')
        else:
            students_for_item = Student.objects.filter(subgroups__in=[subgroup_obj], group=group).order_by(
                'student_full_name')

        lectures = ClassModel.objects.filter(subject__item_name=item, date_class__month=month,
                                             date_class__year=year, group=group, sub_group=subgroup_obj).order_by(
            'date_class')

        student_progresses = StudentsProgress.objects.filter(subject__item_name=item, lesson_date__month=month,
                                                             lesson_date__year=year, class_model__group=group,
                                                             subgroup=subgroup_obj)

        idx = 0
        for lecture in lectures:
            progresses_by_lecture = StudentsProgress.objects.filter(class_model=lecture)
            for progress in progresses_by_lecture:
                if progress and progress.mark_element_id:
                    new_mark_element_id = progress.mark_element_id.split('-')
                    new_mark_element_id[3] = str(idx)
                    string_mark_element_id = '-'.join(new_mark_element_id)
                    progress.mark_element_id = string_mark_element_id
                    progress.save()

            idx += 1

        student_names = []
        student_ids = []
        for student in students_for_item:
            student_names.append(student.student_full_name)
            student_ids.append(student.id)

        lectures_list = []
        for lecture in lectures:
            lectures_list.append({
                "theme": lecture.theme_of_class,
                "type": lecture.type_of_class,
                "date": lecture.date_class.strftime("%Y-%m-%d"),
                "identifier": lecture.identifier,
            })

        progress_list = []
        for single_progress in student_progresses:
            progress_list.append({
                "mark": single_progress.mark,
                "mark_id": single_progress.mark_element_id,
                "id": single_progress.id,
            })

        teachers_by_identifier = {}
        for lecture in lectures:
            teachers_by_identifier[lecture.identifier] = lecture.teacher_for_change

        lesson_types = TypeForLesson.objects.values_list('type_name', flat=True)
        lesson_themes = ThemeForLesson.objects.filter(item=subject)

        lesson_themes_dict = []
        for lesson_theme in lesson_themes:
            lesson_themes_dict.append({lesson_theme.for_type.type_name: lesson_theme.theme_name})

        new_stud_data = {
            "count": len(students_for_item),
            "names": student_names,
            "students": student_ids,
            "lectures_count": len(lectures),
            "lectures": lectures_list,
            "progress": progress_list,
            "month": month,
            "year": year,
            "all_teachers": list(all_teachers),
            "teachers_by_identifier": teachers_by_identifier,
            "lesson_types": list(lesson_types),
            "lesson_themes": list(lesson_themes_dict),
        }

        student_idxs_json = {}
        for lecture in lectures:
            progresses_by_lecture = StudentsProgress.objects.filter(class_model=lecture)
            for progress in progresses_by_lecture:
                student_idxs_json[progress.mark_element_id] = progress.id
                student_idxs_json[f'{progress.mark_element_id}-check'] = not progress.is_valuable
                student_idxs_json[f'{progress.mark_element_id}-check1'] = progress.multi_mark_value

        json_data = json.dumps(new_stud_data)
        student_json_data = json.dumps(student_idxs_json)

        ua_months = [
            'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
            'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень'
        ]

        current_month_name = ua_months[month - 1]
        previous_month = (month - 1) if month > 1 else 12
        next_month = (month + 1) if month < 12 else 1

        operation_history = OperationLog.objects.filter(journal="mark", group=group, subject=subject).order_by(
            '-operation_date')
        not_hidden_history = []
        for operation in operation_history:
            if not request.user in operation.hidden_users.all():
                not_hidden_history.append(operation)

        context = {
            'teachers_for_subject': teachers_for_subject,
            'students_for_item': students_for_item,
            'current_month': month,
            'current_month_name': current_month_name,
            'previous_month': previous_month,
            'previous_month_name': ua_months[previous_month - 1],
            'next_month': next_month,
            'next_month_name': ua_months[next_month - 1],
            'data': json_data,
            'data2': student_json_data,
            'item': item,
            'group': group,
            'curator': curator,
            'history': not_hidden_history,
            'subgroup': subgroup_name,
            "is_journal": True,
        }

        return render(request, 'item_journals/admin_journal.html', context)



class SearchJournalView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, item, group_code, journal, *args, **kwargs):
        journal_urls = {
            "mark": "item_journal",
            "pass": "pass_journal",
            "leave": "leave_journal",
        }

        subgroup_name_list = str(request.META.get('HTTP_REFERER')).split("/")
        subgroup_name = urllib.parse.unquote(subgroup_name_list[len(subgroup_name_list) - 2])

        context = {
            'group_code': group_code,
            'item': item, 'journal': journal,
            'url': journal_urls[journal],
            'subgroup': subgroup_name,
        }

        if "q" in request.GET:
            group = get_object_or_404(CollageGroups, group_code=group_code)
            subject = get_object_or_404(Subject, item_name=item)
            searching_request = request.GET.get('q')

            operation_history = OperationLog.objects.filter(journal="mark", group=group, subject=subject).order_by(
                '-operation_date')
            not_hidden_history = []
            for operation in operation_history:
                if not request.user in operation.hidden_users.all():
                    not_hidden_history.append(operation)

            filtered_ids = []
            for obj in not_hidden_history:
                filtered_ids.append(obj.id)

            searching_results = SearchQuerySet().filter(django_id__in=filtered_ids).autocomplete(
                content_auto=searching_request)

            history = [result.object for result in searching_results]
            context["history"] = history
        else:
            history = []
            context["history"] = history

        return render(request, 'index/history_search_result.html', context)


class ClearHistoryForUserView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, item, group_code, journal, subgroup_name, *args, **kwargs):
        group = get_object_or_404(CollageGroups, group_code=group_code)
        subject = get_object_or_404(Subject, item_name=item)
        operation_history = OperationLog.objects.filter(journal=journal, group=group, subject=subject)
        for operation in operation_history:
            operation.hidden_users.add(request.user)

        journal_urls = {
            "mark": "item_journal",
            "pass": "pass_journal",
            "leave": "leave_journal",
        }

        return redirect(journal_urls[journal], item=item, group_code=group_code, subgroup_name=subgroup_name)


class PassJournalView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, group_code, item, subgroup_name, *args, **kwargs):
        month = int(request.GET.get('month', datetime.now().month))
        year = datetime.now().year

        group = get_object_or_404(CollageGroups, group_code=group_code)
        subject = get_object_or_404(Subject, item_name=item)

        subgroup_obj = Subgroup.objects.filter(name=subgroup_name, parent_group=group).first()
        curator = Curator.objects.filter(group=group).first()
        teachers_for_subject = subject.get_teachers_for_group(group)

        if subgroup_name == "all":
            students_for_item = Student.objects.filter(group=group).order_by(
                'student_full_name')
        else:
            students_for_item = Student.objects.filter(subgroups__in=[subgroup_obj], group=group).order_by(
                'student_full_name')

        lectures = ClassModel.objects.filter(subject__item_name=item, date_class__month=month,
                                             date_class__year=year, group=group, sub_group=subgroup_obj).order_by(
            'date_class')

        student_progresses = StudentsProgress.objects.filter(
            subgroup=subgroup_obj,
            subject__item_name=item,
            lesson_date__month=month,
            lesson_date__year=year
        ).filter(
            Q(mark__in=['Н', 'НБ', 'Г']) |
            Q(workedoutlogs__isnull=False)
        ).distinct()

        idx = 0
        for lecture in lectures:
            progresses_by_lecture = StudentsProgress.objects.filter(class_model=lecture)
            for progress in progresses_by_lecture:
                if progress and progress.mark_element_id:
                    new_mark_element_id = progress.mark_element_id.split('-')
                    new_mark_element_id[3] = str(idx)
                    string_mark_element_id = '-'.join(new_mark_element_id)
                    progress.mark_element_id = string_mark_element_id
                    progress.save()

            idx += 1

        student_names = []
        student_ids = []
        for student in students_for_item:
            student_names.append(student.student_full_name)
            student_ids.append(student.id)

        lectures_list = []
        for lecture in lectures:
            lectures_list.append({
                "theme": lecture.theme_of_class,
                "type": lecture.type_of_class,
                "date": lecture.date_class.strftime("%Y-%m-%d"),
                "identifier": lecture.identifier,
            })

        progress_list = []
        for single_progress in student_progresses:
            progress_list.append({
                "mark": single_progress.mark,
                "mark_id": single_progress.mark_element_id,
                "id": single_progress.id,
            })

        new_stud_data = {
            "count": len(students_for_item),
            "names": student_names,
            "students": student_ids,
            "lectures_count": len(lectures),
            "lectures": lectures_list,
            "progress": progress_list,
            "month": month,
            "year": year,
        }

        student_worked_out = {}
        for progress in student_progresses:
            worked_out = WorkedOutLogs.objects.filter(mark_id=progress).first()
            student_worked_out[progress.mark_element_id] = progress.id
            if worked_out:
                student_worked_out[f'{progress.mark_element_id}-ar'] = worked_out.absent_reason

        json_data = json.dumps(new_stud_data)
        student_json_data = json.dumps(student_worked_out)

        ua_months = [
            'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
            'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень'
        ]

        current_month_name = ua_months[month - 1]
        previous_month = (month - 1) if month > 1 else 12
        next_month = (month + 1) if month < 12 else 1

        operation_history = OperationLog.objects.filter(journal="pass", group=group, subject=subject).order_by(
            '-operation_date')
        not_hidden_history = []
        for operation in operation_history:
            if not request.user in operation.hidden_users.all():
                not_hidden_history.append(operation)

        context = {
            'teachers_for_subject': teachers_for_subject,
            'students_for_item': students_for_item,
            'current_month': month,
            'current_month_name': current_month_name,
            'previous_month': previous_month,
            'previous_month_name': ua_months[previous_month - 1],
            'next_month': next_month,
            'next_month_name': ua_months[next_month - 1],
            'data': json_data,
            'data2': student_json_data,
            'item': item,
            'group': group,
            'history': not_hidden_history,
            'curator': curator,
            'subgroup': subgroup_name,
            "is_journal": True,
        }

        return render(request, 'item_journals/pass_journal_admin.html', context)


class WorkedOutChanges(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                modal_window_data = json.loads(request.body)
                absent_id = modal_window_data.get("absent_id")
                absent_reason = modal_window_data.get("absent_reason")
                worked_on = modal_window_data.get("worked_on")
                before_worked_on = modal_window_data.get("before_worked")

                if not worked_on and not absent_reason:
                    return JsonResponse({"status": "success", "message": "No data"})

                progress_for_update = StudentsProgress.objects.filter(id=absent_id).first()
                lecture = ClassModel.objects.filter(id=progress_for_update.class_model.id).first()
                student = Student.objects.filter(id=progress_for_update.student.id).first()

                if WorkedOutLogs.objects.filter(mark_id=progress_for_update).exists():
                    existing_log = WorkedOutLogs.objects.filter(mark_id=progress_for_update).first()
                    existing_log.absent_reason = absent_reason
                    if worked_on:
                        existing_log.worked_on = worked_on
                        existing_log.is_worked_out = True
                        existing_log.date_worked_out = date.today()
                        existing_log.before_worked_out = before_worked_on
                        progress_for_update.mark = worked_on
                        progress_for_update.save()
                        operation_text = (
                            f'зарахував відпрацювання {lecture.type_of_class}, з темою {lecture.theme_of_class} на {str(lecture.date_class)} '
                            f'для студента {student.student_full_name}.'
                            f'Пропуск відпрацьовано на {worked_on}')

                        log = OperationLog(
                            created_by=request.user,
                            operation=operation_text,
                            operation_date=datetime.now(),
                            journal="pass",
                            subject=lecture.subject,
                            group=lecture.group
                        )
                        log.save()

                    existing_log.save()

                    return JsonResponse({"status": "success", "message": "Successfully updated"})

                worked_out = WorkedOutLogs(
                    absent_reason=absent_reason,
                    mark_id=progress_for_update,
                    before_worked_out=before_worked_on,
                )

                if worked_on:
                    progress_for_update.mark = worked_on
                    progress_for_update.save()
                    worked_out.date_worked_out = date.today()
                    worked_out.worked_on = worked_on
                    worked_out.is_worked_out = True

                operation_text = (
                    f'зарахував відпрацювання {lecture.type_of_class}, з темою {lecture.theme_of_class} на {str(lecture.date_class)} '
                    f'для студента {student.student_full_name}.'
                    f'Пропуск відпрацьовано на {worked_on}')

                log = OperationLog(
                    created_by=request.user,
                    operation=operation_text,
                    operation_date=datetime.now(),
                    journal="pass",
                    subject=lecture.subject,
                    group=lecture.group
                )
                log.save()
                worked_out.save()
                return JsonResponse({"status": "success", "message": "Successfully saved"})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class CuratorJournalView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, *args, **kwargs):
        curator = get_object_or_404(Curator, user=request.user)
        group = curator.group

        month = int(request.GET.get('month', datetime.now().month))
        year = datetime.now().year

        students_in_group = Student.objects.filter(group=group)

        subjects = Subject.objects.filter(
            subjectteachergroup__group=group,
        ).order_by('item_name')

        lectures = ClassModel.objects.filter(
            subject__in=subjects,
            date_class__month=month,
            date_class__year=year,
            group=group
        ).order_by('date_class')

        students_progress = StudentsProgress.objects.filter(
            student__in=students_in_group,
            lesson_date__month=month,
            lesson_date__year=year
        )

        progress_by_student = defaultdict(lambda: defaultdict(list))
        for progress in students_progress:
            progress_by_student[progress.student][progress.subject].append(progress)

        ua_months = [
            'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
            'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень'
        ]

        current_month_name = ua_months[month - 1]
        previous_month = (month - 1) if month > 1 else 12
        next_month = (month + 1) if month < 12 else 1

        context = {
            'current_month': month,
            'current_month_name': current_month_name,
            'previous_month': previous_month,
            'previous_month_name': ua_months[previous_month - 1],
            'next_month': next_month,
            'next_month_name': ua_months[next_month - 1],
            'lectures': lectures,
            'progress_by_student': dict(progress_by_student),
            'group': group,
            'subjects': subjects,
            'students_in_group': students_in_group,
        }

        return render(request, 'item_journals/curator_journal.html', context)


class StudentJournalView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, *args, **kwargs):
        month = int(request.GET.get('month', datetime.now().month))
        year = datetime.now().year

        ua_months = [
            'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
            'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень'
        ]

        current_month_name = ua_months[month - 1]
        previous_month = (month - 1) if month > 1 else 12
        next_month = (month + 1) if month < 12 else 1

        student = Student.objects.filter(user=request.user).first()
        students_progress = StudentsProgress.objects.filter(student=student, lesson_date__month=month,
                                                            lesson_date__year=year)
        subjects = Subject.objects.all()

        progress_by_subject = defaultdict(list)
        for progress in students_progress:
            progress_by_subject[progress.subject].append(progress)

        progress_by_subject = dict(progress_by_subject)

        context = {
            'current_month': month,
            'current_month_name': current_month_name,
            'previous_month': previous_month,
            'previous_month_name': ua_months[previous_month - 1],
            'next_month': next_month,
            'next_month_name': ua_months[next_month - 1],
            "student_progresses": students_progress,
            "student": student,
            "progress_by_subject": progress_by_subject,
        }

        return render(request, 'item_journals/student_journal.html', context)


class PassStudentJournalView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, *args, **kwargs):
        month = int(request.GET.get('month', datetime.now().month))
        year = datetime.now().year

        ua_months = [
            'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
            'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень'
        ]

        current_month_name = ua_months[month - 1]
        previous_month = (month - 1) if month > 1 else 12
        next_month = (month + 1) if month < 12 else 1

        student = Student.objects.filter(user=request.user).first()
        students_progress = StudentsProgress.objects.filter(
            student=student,
            lesson_date__month=month,
            lesson_date__year=year
        ).filter(
            Q(mark__in=['Н', 'НБ', 'Г']) |
            Q(workedoutlogs__isnull=False)
        ).distinct()

        subjects = Subject.objects.all()
        progress_by_subject = defaultdict(list)
        for progress in students_progress:
            worked_out = WorkedOutLogs.objects.filter(mark_id=progress).first()
            progress_dict = {
                'class_model': progress.class_model,
                'lesson_date': progress.lesson_date,
                'mark': progress.mark,
                'worked_out': worked_out if worked_out else None,
            }
            progress_by_subject[progress.subject].append(progress_dict)

        context = {
            'current_month': month,
            'current_month_name': current_month_name,
            'previous_month': previous_month,
            'previous_month_name': ua_months[previous_month - 1],
            'next_month': next_month,
            'next_month_name': ua_months[next_month - 1],
            "student_progresses": students_progress,
            "student": student,
            "progress_by_subject": dict(progress_by_subject),
        }

        return render(request, 'item_journals/pass_journal_student.html', context)


class UpdateMarkView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        mark_data = json.loads(request.body)
        mark_bool = mark_data.get('check_box_result')
        mark_id = mark_data.get('mark_id')
        multi_mark = mark_data.get('multi_check_box_result')

        mark_object = StudentsProgress.objects.filter(id=mark_id).first()
        student = Student.objects.filter(id=mark_object.student.id).first()
        lecture = ClassModel.objects.filter(id=mark_object.class_model.id).first()
        operation_text = (f'відмітив оцінку {mark_object.mark} студенту {student.student_full_name} '
                          f'за {lecture.type_of_class}({str(lecture.date_class)}) з теми {lecture.theme_of_class} як неактивну')
        log = OperationLog(
            created_by=request.user,
            operation=operation_text,
            operation_date=datetime.now(),
            journal="mark",
            subject=lecture.subject,
            group=lecture.group,
        )
        log.save()

        if multi_mark == "":
            mark_object.is_valuable = not mark_bool
            mark_object.save()
        else:
            mark_object.multi_mark_value = multi_mark
            mark_object.save()

        return JsonResponse({'status': 'success', 'message': f'Successfully {mark_bool}'})


class UpdateLectureView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                lecture_data = json.loads(request.body)
                ident = lecture_data.get('ident')
                lecturer_changed = lecture_data.get('lecturer_changed')

                lecture = ClassModel.objects.filter(identifier=ident).first()
                if lecture:
                    lecture.teacher_for_change = lecturer_changed
                    lecture.save()
                    if lecturer_changed != "":
                        operation_text = (
                            f'призначив заміну для заняття з типом {lecture.type_of_class}, з темою {lecture.theme_of_class} на {str(lecture.date_class)}\n'
                            f'заняття провів {lecturer_changed}')
                    else:
                        operation_text = (
                            f'прибрав заміну для заняття з типом {lecture.type_of_class}, з темою {lecture.theme_of_class} на {str(lecture.date_class)}\n')

                    log = OperationLog(
                        created_by=request.user,
                        operation=operation_text,
                        operation_date=datetime.now(),
                        journal="mark",
                        subject=lecture.subject,
                        group=lecture.group,
                    )
                    log.save()

                return JsonResponse({'status': 'success', 'message': f'Successfully'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'{e}'})


class DeleteLectureView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        lecture_data = json.loads(request.body)
        ident = lecture_data.get("lecture_ident")
        lecture = ClassModel.objects.filter(identifier=ident).first()
        operation_text = f'видалив заняття з типом {lecture.type_of_class}, з темою {lecture.theme_of_class} на {str(lecture.date_class)}'
        log = OperationLog(
            created_by=request.user,
            operation=operation_text,
            operation_date=datetime.now(),
            journal="mark",
            subject=lecture.subject,
            group=lecture.group
        )
        log.save()
        ClassModel.objects.filter(identifier=ident).delete()
        return JsonResponse({'status': 'success', 'message': 'Successfully deleted object'})


class DeleteMarkView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        try:
            mark_data = json.loads(request.body)
            mark_id = mark_data.get('mark_id')

            progress = StudentsProgress.objects.filter(id=mark_id).first()
            student = Student.objects.filter(id=progress.student.id).first()
            lecture = ClassModel.objects.filter(id=progress.class_model.id).first()
            operation_text = (f'видалив {progress.mark} студенту {student.student_full_name} '
                              f'за {lecture.type_of_class}({str(lecture.date_class)}) з теми {lecture.theme_of_class}')
            log = OperationLog(
                created_by=request.user,
                operation=operation_text,
                operation_date=datetime.now(),
                journal="mark",
                subject=lecture.subject,
                group=lecture.group,
            )
            log.save()
            StudentsProgress.objects.filter(id=mark_id).delete()

            return JsonResponse({'status': 'success', 'message': f'Successfully {mark_id}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class GetMarkJsonView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        journal_elements = {'Н', 'НБ', 'Г'}
        year = datetime.now().year

        try:
            with transaction.atomic():
                student_data = json.loads(request.body)
                mark = student_data.get('value')
                student_id = student_data.get('student')
                item_name = student_data.get('item_name')
                class_date = student_data.get('class_date')
                random_identifier = student_data.get('identifier')
                mark_element_id = student_data.get('mark_element_id')
                current_month = student_data.get('current_month')
                subgroup = student_data.get('subgroup')
                group_code = student_data.get('group_code')

                decoded_item_name = urllib.parse.unquote(item_name)
                subject = Subject.objects.get(item_name=decoded_item_name)
                student = Student.objects.get(id=student_id)

                decoded_group_code = urllib.parse.unquote(group_code)
                group = CollageGroups.objects.filter(group_code=decoded_group_code).first()

                decoded_subgroup_name = urllib.parse.unquote(subgroup)
                subgroup_obj = Subgroup.objects.filter(name=decoded_subgroup_name, parent_group=group).first()

                lecture = ClassModel.objects.filter(identifier=random_identifier).first()
                if not lecture:
                    return JsonResponse(
                        {'status': 'error', 'message': 'Щоб заповнити оцінки, спочатку створіть заняття'})

                if '/' in mark:
                    mark = str(mark).replace(" ", "")
                    splited_marks = str(mark).split("/")
                    for splited_mark in splited_marks:
                        if splited_mark.isnumeric():
                            if int(splited_mark) > 12 or int(splited_mark) < 1:
                                return JsonResponse({'status': 'error', 'message': 'Введено невалідні дані'})
                elif mark.upper() not in journal_elements or mark.isnumeric():
                    if int(mark) > 12 or int(mark) < 1:
                        return JsonResponse({'status': 'error', 'message': 'Введено невалідні дані'})

                if StudentsProgress.objects.filter(mark_element_id=mark_element_id,
                                                   subject=subject,
                                                   lesson_date__month=current_month,
                                                   lesson_date__year=year, subgroup=subgroup_obj).exists():

                    student_progress = StudentsProgress.objects.filter(mark_element_id=mark_element_id,
                                                                       subject=subject,
                                                                       lesson_date__month=current_month,
                                                                       lesson_date__year=year,
                                                                       subgroup=subgroup_obj).first()
                    operation_text = (
                        f'оновив оцінку з {student_progress.mark} на {mark.upper()}  студенту {student.student_full_name}'
                        f' за {lecture.type_of_class}({str(lecture.date_class)}) з теми {lecture.theme_of_class}')
                    student_progress.mark = mark.upper()
                    student_progress.save()
                    log = OperationLog(
                        created_by=request.user,
                        operation=operation_text,
                        operation_date=datetime.now(),
                        journal="mark",
                        subject=subject,
                        group=lecture.group,
                    )
                    log.save()
                else:
                    student_progress = StudentsProgress(
                        mark=mark.upper(),
                        student=student,
                        subject=subject,
                        class_model=lecture,
                        mark_element_id=mark_element_id,
                        lesson_date=lecture.date_class,
                        subgroup=subgroup_obj,
                    )

                    student_progress.save()
                    operation_text = (f'поставив {mark.upper()} студенту {student.student_full_name} '
                                      f'за {lecture.type_of_class}({str(lecture.date_class)}) з теми {lecture.theme_of_class}')
                    log = OperationLog(
                        created_by=request.user,
                        operation=operation_text,
                        operation_date=datetime.now(),
                        journal="mark",
                        subject=subject,
                        group=lecture.group,
                    )
                    log.save()

                return JsonResponse(
                    {'status': 'success', 'message': f'Data received and saved successfully {year}'})
        except Exception as e:
            print(e)
            return JsonResponse({'status': 'error', 'message': "Введено невалідні дані"})


def generate_random_identifier():
    return ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(10))


class GetLectureJsonView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        try:
            class_data = json.loads(request.body)
            theme = class_data.get('theme')
            class_type = class_data.get('type')
            class_date = class_data.get('date')
            item_name = class_data.get('item')
            ident = class_data.get('identifier')
            group_code = class_data.get('group')
            subgroup_name = class_data.get('subgroup')

            decoded_item_name = urllib.parse.unquote(item_name)
            subject = Subject.objects.get(item_name=decoded_item_name)
            decoded_group = urllib.parse.unquote(group_code)
            decoded_subgroup_name = urllib.parse.unquote(subgroup_name)
            group = CollageGroups.objects.get(group_code=decoded_group)
            subgroup = Subgroup.objects.filter(name=decoded_subgroup_name, parent_group=group).first()

            random_identifier = ident
            if not random_identifier:
                random_identifier = generate_random_identifier()
                while ClassModel.objects.filter(identifier=random_identifier).exists():
                    random_identifier = generate_random_identifier()

                lecture = ClassModel(
                    theme_of_class=theme,
                    type_of_class=class_type,
                    date_class=class_date,
                    subject=subject,
                    group=group,
                    identifier=random_identifier,
                    sub_group=subgroup,
                )

                lecture.save()
                operation_text = f'додав заняття з типом {lecture.type_of_class}, з темою {lecture.theme_of_class} на {str(lecture.date_class)}'
                log = OperationLog(
                    created_by=request.user,
                    operation=operation_text,
                    operation_date=datetime.now(),
                    journal="mark",
                    subject=subject,
                    group=group,

                )
                log.save()
            else:
                lecture = ClassModel.objects.filter(identifier=random_identifier).first()
                operation_text = (f'оновив заняття:\n'
                                  f'    (До оновлення типу){lecture.type_of_class}-(Після оновлення типу){class_type}\n'
                                  f'    (До оновлення теми){lecture.theme_of_class}-(Після оновлення теми){theme}\n'
                                  f'    (До оновлення дати){str(lecture.date_class)}-(Після оновлення дати){str(class_date)}')
                log = OperationLog(
                    created_by=request.user,
                    operation=operation_text,
                    operation_date=datetime.now(),
                    journal="mark",
                    subject=subject,
                    group=group
                )
                log.save()
                lecture.theme_of_class = theme
                lecture.type_of_class = class_type
                lecture.date_class = class_date
                lecture.save()

            return JsonResponse({'status': 'success', 'message': 'Object was created'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class ExportThemesFileTemplateView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Теми для занять"

        columns = ['Назва теми', 'До якого типу належить']
        ws.append(columns)

        column_width_px = 220
        column_width_chars = column_width_px / 7

        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = column_width_chars

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Теми для занять.xlsx'

        wb.save(response)

        return response


class ImportThemesView(View):
    @csrf_exempt
    def post(self, request, *args, **kwargs):
        if 'file' in request.FILES:
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                theme_name = row[0]
                type_theme = row[1]
                prev_url = request.META.get('HTTP_REFERER', '/')

                for item in str(prev_url).split('/'):
                    decoded_item_name = urllib.parse.unquote(urllib.parse.unquote(item))
                    if Subject.objects.filter(item_name=decoded_item_name).exists():
                        subject = Subject.objects.filter(item_name=decoded_item_name).first()

                        lesson_type = TypeForLesson.objects.filter(type_name__iexact=type_theme).first()
                        theme, created = ThemeForLesson.objects.get_or_create(
                            for_type=lesson_type,
                            theme_name=theme_name,
                            item=subject
                        )

        return redirect(request.META.get('HTTP_REFERER', '/'))
